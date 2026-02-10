#!/usr/bin/env python3
"""
DBC Compare Tool
Compares two versions of DBC files and generates color-coded Excel comparison reports.
Covers ALL DBC attributes: message-level and signal-level.

Usage:
    python3 dbc_compare.py <old_folder> <new_folder> [--output <output_folder>]

Color scheme:
    - Orange background (FFA500): Row has differences
    - Red background (FF0000): Specific cell that changed
    - No color: Row is identical between versions

                  _                                        _
  _ __ ___   __ _| | _____   _ __   ___  __ _  ___ ___   | |
 | '_ ` _ \ / _` | |/ / _ \ | '_ \ / _ \/ _` |/ __/ _ \  | |
 | | | | | | (_| |   <  __/ | |_) |  __/ (_| | (_|  __/  |_|
 |_| |_| |_|\__,_|_|\_\___| | .__/ \___|\__,_|\___\___|  (_)
                             |_|
              not war  --  5c1c30200c080e3ff581251ea1dfce8b9e0b12db7b194f8c8a6beb687939f492
   Find yourself in peace: https://www.youtube.com/watch?v=4-079YIasck
"""

import os
import re
import sys
import argparse
from collections import OrderedDict
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# ==============================================================================
# Column Layout Definition
# ==============================================================================

# Message columns (per side)
MSG_HEADERS = [
    "Id",              # 0  - hex message ID
    "Name",            # 1  - message name
    "SendType",        # 2  - GenMsgSendType
    "CycleTime",       # 3  - GenMsgCycleTime
    "CycleTimeFast",   # 4  - GenMsgCycleTimeFast
    "CycleTimeActive", # 5  - GenMsgCycleTimeActive
    "NrOfRepetition",  # 6  - GenMsgNrOfRepetition
    "DelayTime",       # 7  - GenMsgDelayTime
    "StartDelayTime",  # 8  - GenMsgStartDelayTime
    "Dlc",             # 9  - data length code
    "VFrameFormat",    # 10 - CAN / CAN FD
    "CANFD_BRS",       # 11 - bit rate switch
    "Transmitter",     # 12 - transmitting node
    "ILSupport",       # 13 - GenMsgILSupport
    "NmMessage",       # 14 - NM message flag
    "DiagRequest",     # 15 - diagnostic request
    "DiagResponse",    # 16 - diagnostic response
    "DiagState",       # 17 - diagnostic state
    "MsgComment",      # 18 - CM_ BO_
]

# Signal columns (per side)
SIG_HEADERS = [
    "Name",            # 0  - signal name
    "Start",           # 1  - start bit (natural)
    "Length",          # 2  - bit length
    "Endian",          # 3  - big/little
    "Signed",          # 4  - +/-
    "Scale",           # 5  - factor
    "Offset",          # 6  - offset
    "Min",             # 7  - minimum
    "Max",             # 8  - maximum
    "Unit",            # 9  - unit string
    "InvalidValue",    # 10 - BA_ InvalidValue
    "StartValue",      # 11 - GenSigStartValue (init/failsafe)
    "InactiveValue",   # 12 - GenSigInactiveValue
    "SendType",        # 13 - GenSigSendType
    "TimeoutTime",     # 14 - GenSigTimeoutTime_ALL
    "ValueDesc",       # 15 - VAL_ descriptions
    "Comment",         # 16 - CM_ SG_ comment
    "Receivers",       # 17 - receiver nodes
]

NUM_MSG_COLS = len(MSG_HEADERS)
NUM_SIG_COLS = len(SIG_HEADERS)
NUM_COLS_PER_SIDE = NUM_MSG_COLS + NUM_SIG_COLS

# Enum lookup tables
MSG_SEND_TYPE_ENUM = [
    "Cyclic", "NoMsgSendType", "NotUsed", "NotUsed", "NotUsed",
    "NotUsed", "NotUsed", "IfActive", "NoMsgSendType", "NotUsed"
]
SIG_SEND_TYPE_ENUM = [
    "Cyclic", "NoSigSendType", "OnWriteWithRepetition", "OnChange",
    "OnChangeWithRepetition", "IfActive", "IfActiveWithRepetition",
    "NoSigSendType"
]
VFRAME_FORMAT_ENUM = [
    "StandardCAN", "ExtendedCAN", "reserved", "J1939PG",
    "reserved", "reserved", "reserved", "reserved",
    "reserved", "reserved", "reserved", "reserved",
    "reserved", "reserved", "StandardCAN_FD", "ExtendedCAN_FD"
]
CANFD_BRS_ENUM = ["0", "1"]
DIAG_ENUM = ["No", "Yes"]
IL_ENUM = ["No", "Yes"]
NM_ENUM = ["No", "Yes"]


def enum_lookup(enum_list, index, default=""):
    """Safely look up an enum value by index."""
    try:
        idx = int(index)
        if 0 <= idx < len(enum_list):
            return enum_list[idx]
    except (ValueError, TypeError):
        pass
    return default


# ==============================================================================
# DBC Parser
# ==============================================================================

class Signal:
    def __init__(self):
        self.name = ""
        self.start_bit = 0       # DBC raw start bit
        self.length = 0
        self.byte_order = 0      # 0=big endian (Motorola), 1=little endian (Intel)
        self.value_type = "+"    # + unsigned, - signed
        self.factor = 1.0
        self.offset = 0.0
        self.minimum = 0.0
        self.maximum = 0.0
        self.unit = ""
        self.receivers = []
        self.comment = ""        # from CM_ SG_
        self.value_desc = ""     # from VAL_
        # Attributes
        self.invalid_value = ""           # InvalidValue
        self.start_value = 0              # GenSigStartValue
        self.inactive_value = 0           # GenSigInactiveValue
        self.send_type = 0                # GenSigSendType (enum index)
        self.timeout_time = 0             # GenSigTimeoutTime_ALL
        self.long_symbol = ""             # SystemSignalLongSymbol

    @property
    def endian_str(self):
        return "big" if self.byte_order == 0 else "little"

    @property
    def signed_str(self):
        return "signed" if self.value_type == "-" else "unsigned"

    @property
    def natural_start_bit(self):
        """Return the raw DBC start bit position.

        For Motorola (big endian): MSB bit position (standard CANdb++ convention).
        For Intel (little endian): LSB bit position.
        Using the raw DBC value matches Vector CANdb++ and industry-standard tools.
        """
        return self.start_bit

    @property
    def display_name(self):
        """Return long symbol if available and different from name."""
        if self.long_symbol and self.long_symbol != self.name:
            return self.long_symbol
        return self.name


class Message:
    def __init__(self):
        self.msg_id = 0
        self.name = ""
        self.dlc = 0
        self.transmitter = ""
        self.signals = OrderedDict()  # name -> Signal
        self.comment = ""             # from CM_ BO_
        # Attributes
        self.cycle_time = 0           # GenMsgCycleTime
        self.cycle_time_fast = 0      # GenMsgCycleTimeFast
        self.cycle_time_active = 0    # GenMsgCycleTimeActive
        self.send_type = 0            # GenMsgSendType (enum index)
        self.nr_of_repetition = 0     # GenMsgNrOfRepetition
        self.delay_time = 0           # GenMsgDelayTime
        self.start_delay_time = 0     # GenMsgStartDelayTime
        self.il_support = 1           # GenMsgILSupport (default=Yes=1)
        self.nm_message = 0           # NmMessage
        self.nm_asr_message = 0       # NmAsrMessage
        self.vframe_format = 14       # VFrameFormat (default=StandardCAN_FD=14)
        self.canfd_brs = 1            # CANFD_BRS (default=1)
        self.diag_request = 0         # DiagRequest
        self.diag_response = 0        # DiagResponse
        self.diag_state = 0           # DiagState
        self.long_symbol = ""         # SystemMessageLongSymbol

    @property
    def display_name(self):
        if self.long_symbol and self.long_symbol != self.name:
            return self.long_symbol
        return self.name


class DBCDatabase:
    def __init__(self):
        self.messages = OrderedDict()  # msg_id -> Message
        self.nodes = []
        self.bus_type = ""
        self.db_name = ""
        self.baudrate = 500000
        self.baudrate_canfd = 2000000


def parse_dbc(filepath):
    """Parse a DBC file and return a DBCDatabase object with ALL attributes."""
    db = DBCDatabase()

    with open(filepath, 'r', encoding='utf-8', errors='replace') as f:
        content = f.read()

    # Parse nodes (BU_)
    bu_match = re.search(r'^BU_\s*:\s*(.*)', content, re.MULTILINE)
    if bu_match:
        db.nodes = bu_match.group(1).strip().split()

    # Parse messages (BO_) and signals (SG_)
    msg_pattern = re.compile(r'^BO_\s+(\d+)\s+(\w+)\s*:\s*(\d+)\s+(\w+)', re.MULTILINE)
    sig_pattern = re.compile(
        r'^\s+SG_\s+(\w+)\s*:\s*(\d+)\|(\d+)@([01])([+-])\s*\(([^,]+),([^)]+)\)\s*\[([^|]+)\|([^\]]+)\]\s*"([^"]*)"\s*(.*)',
        re.MULTILINE
    )

    msg_matches = list(msg_pattern.finditer(content))
    for i, mm in enumerate(msg_matches):
        msg = Message()
        msg.msg_id = int(mm.group(1))
        msg.name = mm.group(2)
        msg.dlc = int(mm.group(3))
        msg.transmitter = mm.group(4)

        start_pos = mm.end()
        end_pos = msg_matches[i + 1].start() if i + 1 < len(msg_matches) else len(content)
        msg_block = content[start_pos:end_pos]

        for sm in sig_pattern.finditer(msg_block):
            sig = Signal()
            sig.name = sm.group(1)
            sig.start_bit = int(sm.group(2))
            sig.length = int(sm.group(3))
            sig.byte_order = int(sm.group(4))
            sig.value_type = sm.group(5)
            sig.factor = float(sm.group(6))
            sig.offset = float(sm.group(7))
            sig.minimum = float(sm.group(8))
            sig.maximum = float(sm.group(9))
            sig.unit = sm.group(10)
            receivers_str = sm.group(11).strip()
            sig.receivers = [r.strip() for r in receivers_str.split(',') if r.strip()]
            msg.signals[sig.name] = sig

        db.messages[msg.msg_id] = msg

    # ---- Parse ALL message-level attributes (BA_ ... BO_) ----
    def parse_msg_int_attr(attr_name):
        pattern = re.compile(rf'^BA_ "{attr_name}" BO_ (\d+) (-?\d+)\s*;', re.MULTILINE)
        results = {}
        for m in pattern.finditer(content):
            results[int(m.group(1))] = int(m.group(2))
        return results

    # GenMsgCycleTime
    for msg_id, val in parse_msg_int_attr("GenMsgCycleTime").items():
        if msg_id in db.messages: db.messages[msg_id].cycle_time = val
    # GenMsgCycleTimeFast
    for msg_id, val in parse_msg_int_attr("GenMsgCycleTimeFast").items():
        if msg_id in db.messages: db.messages[msg_id].cycle_time_fast = val
    # GenMsgCycleTimeActive
    for msg_id, val in parse_msg_int_attr("GenMsgCycleTimeActive").items():
        if msg_id in db.messages: db.messages[msg_id].cycle_time_active = val
    # GenMsgSendType
    for msg_id, val in parse_msg_int_attr("GenMsgSendType").items():
        if msg_id in db.messages: db.messages[msg_id].send_type = val
    # GenMsgNrOfRepetition
    for msg_id, val in parse_msg_int_attr("GenMsgNrOfRepetition").items():
        if msg_id in db.messages: db.messages[msg_id].nr_of_repetition = val
    # GenMsgDelayTime
    for msg_id, val in parse_msg_int_attr("GenMsgDelayTime").items():
        if msg_id in db.messages: db.messages[msg_id].delay_time = val
    # GenMsgStartDelayTime
    for msg_id, val in parse_msg_int_attr("GenMsgStartDelayTime").items():
        if msg_id in db.messages: db.messages[msg_id].start_delay_time = val
    # GenMsgILSupport
    for msg_id, val in parse_msg_int_attr("GenMsgILSupport").items():
        if msg_id in db.messages: db.messages[msg_id].il_support = val
    # NmMessage
    for msg_id, val in parse_msg_int_attr("NmMessage").items():
        if msg_id in db.messages: db.messages[msg_id].nm_message = val
    # NmAsrMessage
    for msg_id, val in parse_msg_int_attr("NmAsrMessage").items():
        if msg_id in db.messages: db.messages[msg_id].nm_asr_message = val
    # VFrameFormat
    for msg_id, val in parse_msg_int_attr("VFrameFormat").items():
        if msg_id in db.messages: db.messages[msg_id].vframe_format = val
    # CANFD_BRS
    for msg_id, val in parse_msg_int_attr("CANFD_BRS").items():
        if msg_id in db.messages: db.messages[msg_id].canfd_brs = val
    # DiagRequest
    for msg_id, val in parse_msg_int_attr("DiagRequest").items():
        if msg_id in db.messages: db.messages[msg_id].diag_request = val
    # DiagResponse
    for msg_id, val in parse_msg_int_attr("DiagResponse").items():
        if msg_id in db.messages: db.messages[msg_id].diag_response = val
    # DiagState
    for msg_id, val in parse_msg_int_attr("DiagState").items():
        if msg_id in db.messages: db.messages[msg_id].diag_state = val

    # SystemMessageLongSymbol (string attr)
    lm_pattern = re.compile(r'^BA_ "SystemMessageLongSymbol" BO_ (\d+) "([^"]*)"\s*;', re.MULTILINE)
    for m in lm_pattern.finditer(content):
        msg_id = int(m.group(1))
        if msg_id in db.messages:
            db.messages[msg_id].long_symbol = m.group(2)

    # ---- Parse ALL signal-level attributes (BA_ ... SG_) ----
    def parse_sig_int_attr(attr_name):
        pattern = re.compile(rf'^BA_ "{attr_name}" SG_ (\d+) (\w+) (-?\d+)\s*;', re.MULTILINE)
        results = {}
        for m in pattern.finditer(content):
            results[(int(m.group(1)), m.group(2))] = int(m.group(3))
        return results

    def parse_sig_str_attr(attr_name):
        pattern = re.compile(rf'^BA_ "{attr_name}" SG_ (\d+) (\w+) "([^"]*)"\s*;', re.MULTILINE)
        results = {}
        for m in pattern.finditer(content):
            results[(int(m.group(1)), m.group(2))] = m.group(3)
        return results

    def apply_sig_attr(attr_dict, field_name):
        for (msg_id, sig_name), val in attr_dict.items():
            if msg_id in db.messages and sig_name in db.messages[msg_id].signals:
                setattr(db.messages[msg_id].signals[sig_name], field_name, val)

    # GenSigStartValue
    apply_sig_attr(parse_sig_int_attr("GenSigStartValue"), "start_value")
    # GenSigInactiveValue
    apply_sig_attr(parse_sig_int_attr("GenSigInactiveValue"), "inactive_value")
    # GenSigSendType
    apply_sig_attr(parse_sig_int_attr("GenSigSendType"), "send_type")
    # GenSigTimeoutTime_ALL
    apply_sig_attr(parse_sig_int_attr("GenSigTimeoutTime_ALL"), "timeout_time")
    # InvalidValue (string)
    apply_sig_attr(parse_sig_str_attr("InvalidValue"), "invalid_value")
    # SystemSignalLongSymbol (string)
    apply_sig_attr(parse_sig_str_attr("SystemSignalLongSymbol"), "long_symbol")

    # ---- Parse value descriptions (VAL_) ----
    val_pattern = re.compile(r'^VAL_\s+(\d+)\s+(\w+)\s+(.*?)\s*;', re.MULTILINE)
    for vm in val_pattern.finditer(content):
        msg_id = int(vm.group(1))
        sig_name = vm.group(2)
        val_str = vm.group(3).strip()
        if msg_id in db.messages and sig_name in db.messages[msg_id].signals:
            pairs = re.findall(r'(\d+)\s+"([^"]*)"', val_str)
            desc = ", ".join(f"{num}: {desc}" for num, desc in pairs)
            db.messages[msg_id].signals[sig_name].value_desc = desc

    # ---- Parse comments (CM_) ----
    # Signal comments
    cm_sig_pattern = re.compile(r'^CM_\s+SG_\s+(\d+)\s+(\w+)\s+"([^"]*)"\s*;', re.MULTILINE)
    for cm in cm_sig_pattern.finditer(content):
        msg_id = int(cm.group(1))
        sig_name = cm.group(2)
        comment = cm.group(3)
        if msg_id in db.messages and sig_name in db.messages[msg_id].signals:
            db.messages[msg_id].signals[sig_name].comment = comment

    # Message comments
    cm_msg_pattern = re.compile(r'^CM_\s+BO_\s+(\d+)\s+"([^"]*)"\s*;', re.MULTILINE)
    for cm in cm_msg_pattern.finditer(content):
        msg_id = int(cm.group(1))
        comment = cm.group(2)
        if msg_id in db.messages:
            db.messages[msg_id].comment = comment

    # ---- Parse global attributes ----
    bus_match = re.search(r'^BA_\s+"BusType"\s+"([^"]*)"\s*;', content, re.MULTILINE)
    if bus_match:
        db.bus_type = bus_match.group(1)

    name_match = re.search(r'^BA_\s+"DBName"\s+"([^"]*)"\s*;', content, re.MULTILINE)
    if name_match:
        db.db_name = name_match.group(1)

    return db


# ==============================================================================
# Comparison Logic
# ==============================================================================

def format_number(val):
    """Format a number: remove trailing .0 for integers."""
    if isinstance(val, float) and val == int(val):
        return int(val)
    return val


def build_signal_row(msg, sig):
    """Build a full row = message columns + signal columns."""
    msg_id_hex = f"0x{msg.msg_id:X}"
    receivers = ", ".join(sig.receivers)
    val_desc = sig.value_desc if sig.value_desc else None

    msg_part = [
        msg_id_hex,                                              # Id
        msg.display_name,                                        # Name
        enum_lookup(MSG_SEND_TYPE_ENUM, msg.send_type, "Cyclic"),# SendType
        msg.cycle_time,                                          # CycleTime
        msg.cycle_time_fast,                                     # CycleTimeFast
        msg.cycle_time_active,                                   # CycleTimeActive
        msg.nr_of_repetition,                                    # NrOfRepetition
        msg.delay_time,                                          # DelayTime
        msg.start_delay_time,                                    # StartDelayTime
        msg.dlc,                                                 # Dlc
        enum_lookup(VFRAME_FORMAT_ENUM, msg.vframe_format, "StandardCAN_FD"),  # VFrameFormat
        enum_lookup(CANFD_BRS_ENUM, msg.canfd_brs, "1"),        # CANFD_BRS
        msg.transmitter,                                         # Transmitter
        enum_lookup(IL_ENUM, msg.il_support, "Yes"),             # ILSupport
        enum_lookup(NM_ENUM, msg.nm_message, "No"),              # NmMessage
        enum_lookup(DIAG_ENUM, msg.diag_request, "No"),          # DiagRequest
        enum_lookup(DIAG_ENUM, msg.diag_response, "No"),         # DiagResponse
        enum_lookup(DIAG_ENUM, msg.diag_state, "No"),            # DiagState
        msg.comment if msg.comment else None,                    # MsgComment
    ]

    sig_part = [
        sig.display_name,                                        # Name
        sig.natural_start_bit,                                   # Start
        sig.length,                                              # Length
        sig.endian_str,                                          # Endian
        sig.signed_str,                                          # Signed
        format_number(sig.factor),                               # Scale
        format_number(sig.offset),                               # Offset
        format_number(sig.minimum),                              # Min
        format_number(sig.maximum),                              # Max
        sig.unit if sig.unit else None,                          # Unit
        sig.invalid_value if sig.invalid_value else None,        # InvalidValue
        sig.start_value,                                         # StartValue
        sig.inactive_value,                                      # InactiveValue
        enum_lookup(SIG_SEND_TYPE_ENUM, sig.send_type, "Cyclic"),# SendType
        sig.timeout_time,                                        # TimeoutTime
        val_desc,                                                # ValueDesc
        sig.comment if sig.comment else None,                    # Comment
        receivers,                                               # Receivers
    ]

    return msg_part + sig_part


def build_empty_msg_row(msg):
    """Build a row with message info but empty signal fields."""
    msg_id_hex = f"0x{msg.msg_id:X}"
    msg_part = [
        msg_id_hex,
        msg.display_name,
        enum_lookup(MSG_SEND_TYPE_ENUM, msg.send_type, "Cyclic"),
        msg.cycle_time,
        msg.cycle_time_fast,
        msg.cycle_time_active,
        msg.nr_of_repetition,
        msg.delay_time,
        msg.start_delay_time,
        msg.dlc,
        enum_lookup(VFRAME_FORMAT_ENUM, msg.vframe_format, "StandardCAN_FD"),
        enum_lookup(CANFD_BRS_ENUM, msg.canfd_brs, "1"),
        msg.transmitter,
        enum_lookup(IL_ENUM, msg.il_support, "Yes"),
        enum_lookup(NM_ENUM, msg.nm_message, "No"),
        enum_lookup(DIAG_ENUM, msg.diag_request, "No"),
        enum_lookup(DIAG_ENUM, msg.diag_response, "No"),
        enum_lookup(DIAG_ENUM, msg.diag_state, "No"),
        msg.comment if msg.comment else None,
    ]
    sig_part = [None] * NUM_SIG_COLS
    return msg_part + sig_part


def build_empty_row():
    """Build a completely empty row."""
    return [None] * NUM_COLS_PER_SIDE


def compare_dbc_files(old_db, new_db):
    """
    Compare two DBC databases and produce aligned rows.
    Returns list of (old_row, new_row, has_diff, diff_cols) tuples.
    """
    rows = []
    all_msg_ids = sorted(set(list(old_db.messages.keys()) + list(new_db.messages.keys())))

    for msg_id in all_msg_ids:
        old_msg = old_db.messages.get(msg_id)
        new_msg = new_db.messages.get(msg_id)

        if old_msg and new_msg:
            all_sig_names = list(OrderedDict.fromkeys(
                list(old_msg.signals.keys()) + list(new_msg.signals.keys())
            ))

            if not all_sig_names:
                old_row = build_empty_msg_row(old_msg)
                new_row = build_empty_msg_row(new_msg)
                has_diff, diff_cols = check_row_diff(old_row, new_row)
                rows.append((old_row, new_row, has_diff, diff_cols))
                continue

            for sig_name in all_sig_names:
                old_sig = old_msg.signals.get(sig_name)
                new_sig = new_msg.signals.get(sig_name)

                if old_sig and new_sig:
                    old_row = build_signal_row(old_msg, old_sig)
                    new_row = build_signal_row(new_msg, new_sig)
                elif old_sig and not new_sig:
                    old_row = build_signal_row(old_msg, old_sig)
                    new_row = build_empty_msg_row(new_msg)
                elif not old_sig and new_sig:
                    old_row = build_empty_msg_row(old_msg)
                    new_row = build_signal_row(new_msg, new_sig)
                else:
                    continue

                has_diff, diff_cols = check_row_diff(old_row, new_row)
                rows.append((old_row, new_row, has_diff, diff_cols))

        elif old_msg and not new_msg:
            if old_msg.signals:
                for sig_name, sig in old_msg.signals.items():
                    old_row = build_signal_row(old_msg, sig)
                    new_row = build_empty_row()
                    rows.append((old_row, new_row, True, set(range(NUM_COLS_PER_SIDE))))
            else:
                old_row = build_empty_msg_row(old_msg)
                new_row = build_empty_row()
                rows.append((old_row, new_row, True, set(range(NUM_COLS_PER_SIDE))))

        elif not old_msg and new_msg:
            if new_msg.signals:
                for sig_name, sig in new_msg.signals.items():
                    old_row = build_empty_row()
                    new_row = build_signal_row(new_msg, sig)
                    rows.append((old_row, new_row, True, set(range(NUM_COLS_PER_SIDE))))
            else:
                old_row = build_empty_row()
                new_row = build_empty_msg_row(new_msg)
                rows.append((old_row, new_row, True, set(range(NUM_COLS_PER_SIDE))))

    return rows


def check_row_diff(old_row, new_row):
    """Check if two rows differ. Returns (has_diff, diff_col_indices)."""
    diff_cols = set()
    for i in range(len(old_row)):
        if normalize_val(old_row[i]) != normalize_val(new_row[i]):
            diff_cols.add(i)
    return len(diff_cols) > 0, diff_cols


def normalize_val(v):
    """Normalize a value for comparison."""
    if v is None:
        return None
    if isinstance(v, float) and v == int(v):
        return int(v)
    return v


# ==============================================================================
# Excel Output — Style Definitions
# ==============================================================================

# Dark steel-blue header with white bold text
HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
HEADER_FONT = Font(name='Calibri', size=11, color="FFFFFF", bold=True)
# Sub-header row (version labels) — lighter blue with dark text
SUBHEADER_FILL = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
SUBHEADER_FONT = Font(name='Calibri', size=10, color="1F4E79", bold=True)
# Changed cell — soft pink/red
DIFF_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
# Context row (has diff somewhere but this cell unchanged) — light green tint
CONTEXT_FILL = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
# Alternating row — light gray
ALT_ROW_FILL = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
# Standard data font
DATA_FONT = Font(name='Calibri', size=10, color="333333")
# Title / heading styles for summary sheet
TITLE_FONT = Font(name='Calibri', size=16, color="1F4E79", bold=True)
SUBTITLE_FONT = Font(name='Calibri', size=11, color="555555", italic=True)
METRIC_FONT = Font(name='Calibri', size=11, color="333333")
METRIC_VAL_FONT = Font(name='Calibri', size=12, color="1F4E79", bold=True)
# Badge fills for metric categories
BADGE_NEW = PatternFill(start_color="D5F5E3", end_color="D5F5E3", fill_type="solid")
BADGE_REMOVED = PatternFill(start_color="FADBD8", end_color="FADBD8", fill_type="solid")
BADGE_MODIFIED = PatternFill(start_color="FEF9E7", end_color="FEF9E7", fill_type="solid")
BADGE_TOTAL = PatternFill(start_color="EBF5FB", end_color="EBF5FB", fill_type="solid")
# Thin border for all data cells
THIN_BORDER = Border(
    left=Side(style='thin', color='D5D8DC'),
    right=Side(style='thin', color='D5D8DC'),
    top=Side(style='thin', color='D5D8DC'),
    bottom=Side(style='thin', color='D5D8DC'),
)
# Legacy aliases kept for backward compat within the full comparison sheet
ORANGE_FILL = CONTEXT_FILL
RED_FILL = DIFF_FILL
BLACK_FONT = DATA_FONT


def get_msg_rx_ecus(msg):
    """Get all receiver ECU names for a message (union across all signals)."""
    rx = set()
    for sig in msg.signals.values():
        for r in sig.receivers:
            if r and r != 'Vector__XXX':
                rx.add(r)
    return ', '.join(sorted(rx))


def get_msg_signal_list(msg):
    """Get newline-separated signal names for a message."""
    return '\n'.join(sig.display_name for sig in msg.signals.values())


def build_msg_summary_row(msg):
    """Build a summary row for new/removed messages sheets (8 columns)."""
    return [
        f"0x{msg.msg_id:X}",
        msg.display_name,
        msg.dlc,
        get_msg_signal_list(msg),
        enum_lookup(MSG_SEND_TYPE_ENUM, msg.send_type, "Cyclic"),
        msg.cycle_time,
        msg.transmitter,
        get_msg_rx_ecus(msg),
    ]


def build_sig_summary_row(msg, sig):
    """Build a summary row for new/removed signals sheets (12 columns)."""
    return [
        f"0x{msg.msg_id:X}",
        msg.display_name,
        sig.display_name,
        sig.natural_start_bit,
        sig.length,
        sig.start_value,
        format_number(sig.minimum),
        format_number(sig.maximum),
        sig.unit if sig.unit else "",
        format_number(sig.factor),
        sig.value_desc if sig.value_desc else "",
        ', '.join(sig.receivers),
    ]


def categorize_changes(old_db, new_db):
    """Categorize changes into new/removed/modified messages and signals."""
    new_messages = []
    removed_messages = []
    modified_messages = []
    new_signals = []
    removed_signals = []
    modified_signals = []

    all_msg_ids = sorted(set(list(old_db.messages.keys()) + list(new_db.messages.keys())))

    for msg_id in all_msg_ids:
        old_msg = old_db.messages.get(msg_id)
        new_msg = new_db.messages.get(msg_id)

        if old_msg and not new_msg:
            removed_messages.append(old_msg)
            for sig in old_msg.signals.values():
                removed_signals.append((old_msg, sig))

        elif new_msg and not old_msg:
            new_messages.append(new_msg)
            for sig in new_msg.signals.values():
                new_signals.append((new_msg, sig))

        else:
            # Check message-level changes
            old_row = build_msg_summary_row(old_msg)
            new_row = build_msg_summary_row(new_msg)
            msg_diffs = [i for i in range(len(old_row))
                         if normalize_val(old_row[i]) != normalize_val(new_row[i])]
            if msg_diffs:
                modified_messages.append((old_msg, new_msg, msg_diffs))

            # Check signal-level changes
            all_sigs = list(OrderedDict.fromkeys(
                list(old_msg.signals.keys()) + list(new_msg.signals.keys())
            ))
            for sig_name in all_sigs:
                old_sig = old_msg.signals.get(sig_name)
                new_sig = new_msg.signals.get(sig_name)

                if old_sig and not new_sig:
                    removed_signals.append((old_msg, old_sig))
                elif new_sig and not old_sig:
                    new_signals.append((new_msg, new_sig))
                elif old_sig and new_sig:
                    old_sr = build_sig_summary_row(old_msg, old_sig)[3:]  # skip msg id/name/sig name
                    new_sr = build_sig_summary_row(new_msg, new_sig)[3:]
                    sig_diffs = [i for i in range(len(old_sr))
                                 if normalize_val(old_sr[i]) != normalize_val(new_sr[i])]
                    if sig_diffs:
                        modified_signals.append(
                            (old_msg, new_msg, old_sig, new_sig, sig_diffs))

    return {
        'new_messages': new_messages,
        'removed_messages': removed_messages,
        'modified_messages': modified_messages,
        'new_signals': new_signals,
        'removed_signals': removed_signals,
        'modified_signals': modified_signals,
    }


# Headers for categorized sheets
CAT_MSG_HEADERS = [
    "Message ID", "Message Name", "Message Length", "Signals",
    "Send Type", "Cycle Time", "Tx ECU", "Rx ECUs"
]
CAT_SIG_HEADERS = [
    "Message ID", "Message Name", "Signal Name", "Start Bit", "Length",
    "Initial", "Minimum", "Maximum", "Unit", "Factor", "Value Table", "Receivers"
]
# Paired fields in Modified Signals (after the 3 fixed columns)
MOD_SIG_PAIRED_HEADERS = [
    "Start Bit", "Length", "Initial", "Minimum", "Maximum",
    "Unit", "Factor", "Value Table", "Rx ECUs"
]


def _style_header_row(ws, row, col_start, col_end):
    """Apply header styling (dark fill + white bold) to a row range."""
    for c in range(col_start, col_end + 1):
        cell = ws.cell(row, c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal='center', vertical='center')


def _style_data_cell(ws, row, col, value, alt=False):
    """Write a data cell with border, font, and optional alternating shade."""
    cell = ws.cell(row, col, value)
    cell.font = DATA_FONT
    cell.border = THIN_BORDER
    if alt:
        cell.fill = ALT_ROW_FILL
    return cell


def _write_summary_sheet(wb, old_db, new_db, cats, old_label, new_label):
    """Write the Comparison Summary sheet with a redesigned dashboard layout."""
    ws = wb.create_sheet("Comparison Summary", 0)
    old_msg_count = len(old_db.messages)
    new_msg_count = len(new_db.messages)

    # ── Title block ──
    ws.merge_cells('A1:D1')
    c = ws.cell(1, 1, "DBC Comparison Report")
    c.font = TITLE_FONT
    c.alignment = Alignment(horizontal='left', vertical='center')
    ws.row_dimensions[1].height = 30

    ws.merge_cells('A2:D2')
    c = ws.cell(2, 1, f"{old_label}  vs  {new_label}")
    c.font = SUBTITLE_FONT

    # ── Section: Messages overview ──
    r = 4
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
    ws.cell(r, 1, "Messages Overview").font = Font(name='Calibri', size=12, color="1F4E79", bold=True)
    ws.cell(r, 1).fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
    for c in range(1, 5):
        ws.cell(r, c).border = THIN_BORDER

    metrics_msgs = [
        (f"Total in {old_label}", old_msg_count, BADGE_TOTAL),
        (f"Total in {new_label}", new_msg_count, BADGE_TOTAL),
        ("New Messages", len(cats['new_messages']), BADGE_NEW),
        ("Removed Messages", len(cats['removed_messages']), BADGE_REMOVED),
        ("Modified Messages", len(cats['modified_messages']), BADGE_MODIFIED),
    ]
    for i, (label, val, badge) in enumerate(metrics_msgs):
        row = r + 1 + i
        c_label = ws.cell(row, 1, label)
        c_label.font = METRIC_FONT
        c_label.border = THIN_BORDER
        c_val = ws.cell(row, 2, val)
        c_val.font = METRIC_VAL_FONT
        c_val.border = THIN_BORDER
        c_val.alignment = Alignment(horizontal='center')
        c_val.fill = badge
        ws.cell(row, 3).border = THIN_BORDER
        ws.cell(row, 4).border = THIN_BORDER

    # ── Section: Signals overview ──
    r2 = r + len(metrics_msgs) + 2
    ws.merge_cells(start_row=r2, start_column=1, end_row=r2, end_column=4)
    ws.cell(r2, 1, "Signals Overview").font = Font(name='Calibri', size=12, color="1F4E79", bold=True)
    ws.cell(r2, 1).fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
    for c in range(1, 5):
        ws.cell(r2, c).border = THIN_BORDER

    old_sig_count = sum(len(m.signals) for m in old_db.messages.values())
    new_sig_count = sum(len(m.signals) for m in new_db.messages.values())
    metrics_sigs = [
        (f"Total in {old_label}", old_sig_count, BADGE_TOTAL),
        (f"Total in {new_label}", new_sig_count, BADGE_TOTAL),
        ("New Signals", len(cats['new_signals']), BADGE_NEW),
        ("Removed Signals", len(cats['removed_signals']), BADGE_REMOVED),
        ("Modified Signals", len(cats['modified_signals']), BADGE_MODIFIED),
    ]
    for i, (label, val, badge) in enumerate(metrics_sigs):
        row = r2 + 1 + i
        c_label = ws.cell(row, 1, label)
        c_label.font = METRIC_FONT
        c_label.border = THIN_BORDER
        c_val = ws.cell(row, 2, val)
        c_val.font = METRIC_VAL_FONT
        c_val.border = THIN_BORDER
        c_val.alignment = Alignment(horizontal='center')
        c_val.fill = badge
        ws.cell(row, 3).border = THIN_BORDER
        ws.cell(row, 4).border = THIN_BORDER

    ws.column_dimensions['A'].width = 32
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 5
    ws.column_dimensions['D'].width = 5


def _write_new_removed_msg_sheet(wb, title, messages):
    """Write New Messages or Removed Messages sheet."""
    ws = wb.create_sheet(title)
    ncols = len(CAT_MSG_HEADERS)
    _style_header_row(ws, 1, 1, ncols)
    for i, h in enumerate(CAT_MSG_HEADERS, 1):
        ws.cell(1, i, h)
    for r_idx, msg in enumerate(messages, 2):
        row = build_msg_summary_row(msg)
        alt = (r_idx % 2 == 0)
        for c_idx, val in enumerate(row, 1):
            _style_data_cell(ws, r_idx, c_idx, val, alt=alt)
    for i, w in enumerate([12, 30, 14, 40, 14, 12, 15, 40], 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"


def _write_modified_msg_sheet(wb, modified_messages, old_label, new_label):
    """Write Modified Messages sheet with side-by-side paired columns."""
    ws = wb.create_sheet("Modified Messages")
    num_fields = len(CAT_MSG_HEADERS)
    total_cols = num_fields * 2

    # Row 1: merged field headers
    for i, h in enumerate(CAT_MSG_HEADERS):
        col_start = i * 2 + 1
        ws.merge_cells(start_row=1, start_column=col_start,
                       end_row=1, end_column=col_start + 1)
        c = ws.cell(1, col_start, h)
        c.fill = HEADER_FILL
        c.font = HEADER_FONT
        c.border = THIN_BORDER
        c.alignment = Alignment(horizontal='center')
        ws.cell(1, col_start + 1).fill = HEADER_FILL
        ws.cell(1, col_start + 1).border = THIN_BORDER

    # Row 2: version sub-headers
    for i in range(num_fields):
        col_old = i * 2 + 1
        col_new = i * 2 + 2
        for col, lbl in [(col_old, old_label), (col_new, new_label)]:
            c = ws.cell(2, col, lbl)
            c.font = SUBHEADER_FONT
            c.fill = SUBHEADER_FILL
            c.border = THIN_BORDER
            c.alignment = Alignment(horizontal='center')

    # Data rows
    for r_idx, (old_msg, new_msg, diff_indices) in enumerate(modified_messages, 3):
        old_row = build_msg_summary_row(old_msg)
        new_row = build_msg_summary_row(new_msg)
        alt = (r_idx % 2 == 1)
        for i in range(num_fields):
            c_old = _style_data_cell(ws, r_idx, i * 2 + 1, old_row[i], alt=alt)
            c_new = _style_data_cell(ws, r_idx, i * 2 + 2, new_row[i], alt=alt)
            if i in diff_indices:
                c_old.fill = DIFF_FILL
                c_new.fill = DIFF_FILL

    for i, w in enumerate([12, 12, 30, 30, 14, 14, 40, 40, 14, 14, 12, 12, 15, 15, 40, 40], 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A3"


def _write_new_removed_sig_sheet(wb, title, signal_list):
    """Write New Signals or Removed Signals sheet."""
    ws = wb.create_sheet(title)
    ncols = len(CAT_SIG_HEADERS)
    _style_header_row(ws, 1, 1, ncols)
    for i, h in enumerate(CAT_SIG_HEADERS, 1):
        ws.cell(1, i, h)
    for r_idx, (msg, sig) in enumerate(signal_list, 2):
        row = build_sig_summary_row(msg, sig)
        alt = (r_idx % 2 == 0)
        for c_idx, val in enumerate(row, 1):
            _style_data_cell(ws, r_idx, c_idx, val, alt=alt)
    for i, w in enumerate([12, 30, 35, 10, 8, 10, 10, 10, 10, 10, 40, 30], 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"


def _write_modified_sig_sheet(wb, modified_signals, old_label, new_label):
    """Write Modified Signals sheet with 3 fixed columns + paired fields."""
    ws = wb.create_sheet("Modified Signals")
    fixed_headers = ["Message ID", "Message Name", "Signal Name"]
    num_fixed = len(fixed_headers)
    num_paired = len(MOD_SIG_PAIRED_HEADERS)

    # Row 1: fixed columns merged across rows 1-2, paired columns merged across 2 cols
    for i, h in enumerate(fixed_headers):
        ws.merge_cells(start_row=1, start_column=i + 1, end_row=2, end_column=i + 1)
        c = ws.cell(1, i + 1, h)
        c.fill = HEADER_FILL
        c.font = HEADER_FONT
        c.border = THIN_BORDER
        c.alignment = Alignment(horizontal='center', vertical='center')
    for i, h in enumerate(MOD_SIG_PAIRED_HEADERS):
        col_start = num_fixed + i * 2 + 1
        ws.merge_cells(start_row=1, start_column=col_start,
                       end_row=1, end_column=col_start + 1)
        c = ws.cell(1, col_start, h)
        c.fill = HEADER_FILL
        c.font = HEADER_FONT
        c.border = THIN_BORDER
        c.alignment = Alignment(horizontal='center')
        ws.cell(1, col_start + 1).fill = HEADER_FILL
        ws.cell(1, col_start + 1).border = THIN_BORDER
        for col, lbl in [(col_start, old_label), (col_start + 1, new_label)]:
            c2 = ws.cell(2, col, lbl)
            c2.font = SUBHEADER_FONT
            c2.fill = SUBHEADER_FILL
            c2.border = THIN_BORDER
            c2.alignment = Alignment(horizontal='center')

    # Data rows
    for r_idx, (old_msg, new_msg, old_sig, new_sig, diff_indices) in enumerate(
            modified_signals, 3):
        alt = (r_idx % 2 == 1)
        _style_data_cell(ws, r_idx, 1, f"0x{old_msg.msg_id:X}", alt=alt)
        _style_data_cell(ws, r_idx, 2, old_msg.display_name, alt=alt)
        _style_data_cell(ws, r_idx, 3, old_sig.display_name, alt=alt)

        old_vals = [
            old_sig.natural_start_bit, old_sig.length, old_sig.start_value,
            format_number(old_sig.minimum), format_number(old_sig.maximum),
            old_sig.unit, format_number(old_sig.factor),
            old_sig.value_desc if old_sig.value_desc else "",
            ', '.join(old_sig.receivers),
        ]
        new_vals = [
            new_sig.natural_start_bit, new_sig.length, new_sig.start_value,
            format_number(new_sig.minimum), format_number(new_sig.maximum),
            new_sig.unit, format_number(new_sig.factor),
            new_sig.value_desc if new_sig.value_desc else "",
            ', '.join(new_sig.receivers),
        ]

        for i in range(num_paired):
            c_old = _style_data_cell(ws, r_idx, num_fixed + i * 2 + 1, old_vals[i], alt=alt)
            c_new = _style_data_cell(ws, r_idx, num_fixed + i * 2 + 2, new_vals[i], alt=alt)
            if i in diff_indices:
                c_old.fill = DIFF_FILL
                c_new.fill = DIFF_FILL

    widths = [12, 30, 35] + [12, 12] * num_paired
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A3"


def write_comparison_xlsx(filepath, old_filename, new_filename, comparison_rows,
                          old_db=None, new_db=None, old_label="Old", new_label="New"):
    """Write comparison results to an xlsx file with optional categorized sheets."""
    wb = Workbook()
    wb.properties.creator = "5c1c30200c080e3ff581251ea1dfce8b9e0b12db7b194f8c8a6beb687939f492"
    wb.properties.description = "Make peace, not war. Find yourself in peace: https://www.youtube.com/watch?v=4-079YIasck"

    # --- Categorized sheets (added first so they appear before the flat sheet) ---
    if old_db and new_db:
        cats = categorize_changes(old_db, new_db)
        _write_summary_sheet(wb, old_db, new_db, cats, old_label, new_label)
        _write_new_removed_msg_sheet(wb, "New Messages", cats['new_messages'])
        _write_new_removed_msg_sheet(wb, "Removed Messages", cats['removed_messages'])
        _write_modified_msg_sheet(wb, cats['modified_messages'], old_label, new_label)
        _write_new_removed_sig_sheet(wb, "New Signals", cats['new_signals'])
        _write_new_removed_sig_sheet(wb, "Removed Signals", cats['removed_signals'])
        _write_modified_sig_sheet(wb, cats['modified_signals'], old_label, new_label)

    # --- Full Comparison sheet (flat side-by-side) ---
    ws = wb.create_sheet("Full Comparison") if old_db and new_db else wb.active
    if not (old_db and new_db):
        ws.title = "Sheet1"

    total_cols = NUM_COLS_PER_SIDE * 2
    right_start = NUM_COLS_PER_SIDE + 1  # 1-based column for right side

    # Helper: get column letter
    def col_letter(col_1based):
        return get_column_letter(col_1based)

    last_col_letter = col_letter(total_cols)

    # --- Row 1: File names (dark header) ---
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=NUM_COLS_PER_SIDE)
    ws.merge_cells(start_row=1, start_column=right_start, end_row=1, end_column=total_cols)
    for c in range(1, total_cols + 1):
        ws.cell(1, c).fill = HEADER_FILL
        ws.cell(1, c).border = THIN_BORDER
    ws.cell(1, 1, old_filename).font = HEADER_FONT
    ws.cell(1, right_start, new_filename).font = HEADER_FONT

    # --- Row 2: Section headers (Message | Signal) — sub-header style ---
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=NUM_MSG_COLS)
    ws.merge_cells(start_row=2, start_column=NUM_MSG_COLS + 1, end_row=2, end_column=NUM_COLS_PER_SIDE)
    ws.merge_cells(start_row=2, start_column=right_start, end_row=2, end_column=right_start + NUM_MSG_COLS - 1)
    ws.merge_cells(start_row=2, start_column=right_start + NUM_MSG_COLS, end_row=2, end_column=total_cols)
    for c in range(1, total_cols + 1):
        ws.cell(2, c).fill = SUBHEADER_FILL
        ws.cell(2, c).border = THIN_BORDER
    ws.cell(2, 1, "Message").font = SUBHEADER_FONT
    ws.cell(2, NUM_MSG_COLS + 1, "Signal").font = SUBHEADER_FONT
    ws.cell(2, right_start, "Message").font = SUBHEADER_FONT
    ws.cell(2, right_start + NUM_MSG_COLS, "Signal").font = SUBHEADER_FONT

    # --- Row 3: Column headers (dark header) ---
    all_headers = MSG_HEADERS + SIG_HEADERS
    _style_header_row(ws, 3, 1, total_cols)
    for i, h in enumerate(all_headers):
        ws.cell(3, i + 1, h)
        ws.cell(3, right_start + i, h)

    # --- Data rows ---
    for row_idx, (old_row, new_row, has_diff, diff_cols) in enumerate(comparison_rows, start=4):
        alt = (row_idx % 2 == 0)
        for col_idx in range(NUM_COLS_PER_SIDE):
            cell_old = _style_data_cell(ws, row_idx, col_idx + 1, old_row[col_idx], alt=alt)
            cell_new = _style_data_cell(ws, row_idx, right_start + col_idx, new_row[col_idx], alt=alt)

            if has_diff:
                if col_idx in diff_cols:
                    cell_old.fill = DIFF_FILL
                    cell_new.fill = DIFF_FILL
                else:
                    cell_old.fill = CONTEXT_FILL
                    cell_new.fill = CONTEXT_FILL

    # --- Formatting ---
    ws.freeze_panes = "A4"
    ws.auto_filter.ref = f"A3:{last_col_letter}3"

    # Column widths for both sides
    def set_col_widths(start_col):
        """Set column widths starting from start_col (1-based)."""
        msg_widths = [8, 30, 12, 10, 12, 12, 12, 10, 12, 6, 16, 10, 15, 10, 10, 12, 12, 10, 30]
        sig_widths = [35, 7, 8, 8, 10, 8, 8, 8, 8, 12, 14, 12, 14, 14, 12, 50, 35, 40]
        all_widths = msg_widths + sig_widths
        for i, w in enumerate(all_widths):
            ws.column_dimensions[col_letter(start_col + i)].width = w

    set_col_widths(1)
    set_col_widths(right_start)

    # Remove the default empty "Sheet" created by Workbook() when categorized sheets exist
    if old_db and new_db and "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    wb.save(filepath)
    print(f"  Saved: {filepath}")


# ==============================================================================
# HTML Report
# ==============================================================================

def _html_table(headers, rows, diff_col_indices=None, paired=False):
    """Build an HTML table string. diff_col_indices is a list of sets per row."""
    h = '<table>\n<thead><tr>'
    for hdr in headers:
        h += f'<th>{hdr}</th>'
    h += '</tr></thead>\n<tbody>\n'
    for r_idx, row in enumerate(rows):
        cls = ' class="alt"' if r_idx % 2 == 1 else ''
        h += f'<tr{cls}>'
        diffs = diff_col_indices[r_idx] if diff_col_indices else set()
        for c_idx, val in enumerate(row):
            td_cls = ' class="diff"' if c_idx in diffs else ''
            display = str(val) if val is not None and val != '' else '&mdash;'
            display = display.replace('\n', '<br>')
            h += f'<td{td_cls}>{display}</td>'
        h += '</tr>\n'
    h += '</tbody>\n</table>\n'
    return h


def write_html_report(filepath, old_label, new_label, cats, old_db, new_db):
    """Generate a standalone HTML comparison report."""
    from datetime import datetime
    old_msg_count = len(old_db.messages)
    new_msg_count = len(new_db.messages)
    old_sig_count = sum(len(m.signals) for m in old_db.messages.values())
    new_sig_count = sum(len(m.signals) for m in new_db.messages.values())

    css = """
    * { margin: 0; padding: 0; box-sizing: border-box; }
    body { font-family: 'Segoe UI', Tahoma, sans-serif; color: #333; background: #f4f6f9; padding: 30px; }
    .container { max-width: 1400px; margin: 0 auto; }
    h1 { color: #1F4E79; font-size: 28px; margin-bottom: 4px; }
    .subtitle { color: #777; font-size: 14px; margin-bottom: 20px; }
    .cards { display: flex; gap: 16px; flex-wrap: wrap; margin-bottom: 30px; }
    .card { background: #fff; border-radius: 8px; padding: 18px 24px; box-shadow: 0 2px 8px rgba(0,0,0,0.07); min-width: 150px; flex: 1; }
    .card .label { font-size: 12px; color: #888; text-transform: uppercase; letter-spacing: 0.5px; }
    .card .value { font-size: 28px; font-weight: 700; color: #1F4E79; margin-top: 4px; }
    .card.new .value { color: #27AE60; }
    .card.removed .value { color: #E74C3C; }
    .card.modified .value { color: #F39C12; }
    section { background: #fff; border-radius: 8px; padding: 24px; margin-bottom: 24px; box-shadow: 0 2px 8px rgba(0,0,0,0.07); }
    section h2 { color: #1F4E79; font-size: 18px; margin-bottom: 14px; border-bottom: 2px solid #D6E4F0; padding-bottom: 8px; }
    table { width: 100%; border-collapse: collapse; font-size: 13px; }
    th { background: #1F4E79; color: #fff; padding: 10px 8px; text-align: left; font-weight: 600; }
    td { padding: 8px; border-bottom: 1px solid #e8e8e8; }
    tr.alt td { background: #f8f9fa; }
    td.diff { background: #FFC7CE !important; font-weight: 600; }
    .empty { color: #aaa; font-style: italic; padding: 20px 0; }
    .footer { text-align: center; color: #aaa; font-size: 11px; margin-top: 30px; }
    """

    parts = [f"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>DBC Comparison Report</title>
<style>{css}</style>
</head>
<body>
<div class="container">
<h1>DBC Comparison Report</h1>
<p class="subtitle">{old_label} &rarr; {new_label} &bull; Generated {datetime.now().strftime('%Y-%m-%d %H:%M')}</p>

<div class="cards">
<div class="card"><div class="label">Messages (Base)</div><div class="value">{old_msg_count}</div></div>
<div class="card"><div class="label">Messages (New)</div><div class="value">{new_msg_count}</div></div>
<div class="card"><div class="label">Signals (Base)</div><div class="value">{old_sig_count}</div></div>
<div class="card"><div class="label">Signals (New)</div><div class="value">{new_sig_count}</div></div>
</div>

<div class="cards">
<div class="card new"><div class="label">New Messages</div><div class="value">{len(cats['new_messages'])}</div></div>
<div class="card removed"><div class="label">Removed Messages</div><div class="value">{len(cats['removed_messages'])}</div></div>
<div class="card modified"><div class="label">Modified Messages</div><div class="value">{len(cats['modified_messages'])}</div></div>
<div class="card new"><div class="label">New Signals</div><div class="value">{len(cats['new_signals'])}</div></div>
<div class="card removed"><div class="label">Removed Signals</div><div class="value">{len(cats['removed_signals'])}</div></div>
<div class="card modified"><div class="label">Modified Signals</div><div class="value">{len(cats['modified_signals'])}</div></div>
</div>
"""]

    # New Messages
    parts.append('<section><h2>New Messages</h2>\n')
    if cats['new_messages']:
        rows = [build_msg_summary_row(m) for m in cats['new_messages']]
        parts.append(_html_table(CAT_MSG_HEADERS, rows))
    else:
        parts.append('<p class="empty">No new messages</p>')
    parts.append('</section>\n')

    # Removed Messages
    parts.append('<section><h2>Removed Messages</h2>\n')
    if cats['removed_messages']:
        rows = [build_msg_summary_row(m) for m in cats['removed_messages']]
        parts.append(_html_table(CAT_MSG_HEADERS, rows))
    else:
        parts.append('<p class="empty">No removed messages</p>')
    parts.append('</section>\n')

    # Modified Messages
    parts.append('<section><h2>Modified Messages</h2>\n')
    if cats['modified_messages']:
        paired_hdrs = []
        for h in CAT_MSG_HEADERS:
            paired_hdrs.extend([f"{h} (Base)", f"{h} (New)"])
        rows = []
        diffs_list = []
        for old_msg, new_msg, diff_indices in cats['modified_messages']:
            old_row = build_msg_summary_row(old_msg)
            new_row = build_msg_summary_row(new_msg)
            paired_row = []
            diff_set = set()
            for i in range(len(old_row)):
                paired_row.extend([old_row[i], new_row[i]])
                if i in diff_indices:
                    diff_set.add(i * 2)
                    diff_set.add(i * 2 + 1)
            rows.append(paired_row)
            diffs_list.append(diff_set)
        parts.append(_html_table(paired_hdrs, rows, diffs_list))
    else:
        parts.append('<p class="empty">No modified messages</p>')
    parts.append('</section>\n')

    # New Signals
    parts.append('<section><h2>New Signals</h2>\n')
    if cats['new_signals']:
        rows = [build_sig_summary_row(m, s) for m, s in cats['new_signals']]
        parts.append(_html_table(CAT_SIG_HEADERS, rows))
    else:
        parts.append('<p class="empty">No new signals</p>')
    parts.append('</section>\n')

    # Removed Signals
    parts.append('<section><h2>Removed Signals</h2>\n')
    if cats['removed_signals']:
        rows = [build_sig_summary_row(m, s) for m, s in cats['removed_signals']]
        parts.append(_html_table(CAT_SIG_HEADERS, rows))
    else:
        parts.append('<p class="empty">No removed signals</p>')
    parts.append('</section>\n')

    # Modified Signals
    parts.append('<section><h2>Modified Signals</h2>\n')
    if cats['modified_signals']:
        fixed = ["Msg ID", "Msg Name", "Signal"]
        paired_hdrs = list(fixed)
        for h in MOD_SIG_PAIRED_HEADERS:
            paired_hdrs.extend([f"{h} (Base)", f"{h} (New)"])
        rows = []
        diffs_list = []
        for old_msg, new_msg, old_sig, new_sig, diff_indices in cats['modified_signals']:
            old_sr = build_sig_summary_row(old_msg, old_sig)
            new_sr = build_sig_summary_row(new_msg, new_sig)
            row_data = [old_sr[0], old_sr[1], old_sr[2]]  # fixed cols
            diff_set = set()
            for i in range(len(MOD_SIG_PAIRED_HEADERS)):
                row_data.extend([old_sr[3 + i], new_sr[3 + i]])
                if i in diff_indices:
                    diff_set.add(len(fixed) + i * 2)
                    diff_set.add(len(fixed) + i * 2 + 1)
            rows.append(row_data)
            diffs_list.append(diff_set)
        parts.append(_html_table(paired_hdrs, rows, diffs_list))
    else:
        parts.append('<p class="empty">No modified signals</p>')
    parts.append('</section>\n')

    parts.append('<div class="footer">Generated by DBC Compare Tool</div>\n')
    parts.append('</div></body></html>')

    with open(filepath, 'w', encoding='utf-8') as f:
        f.write(''.join(parts))
    print(f"  Saved HTML: {filepath}")


# ==============================================================================
# PDF Report
# ==============================================================================

def _pdf_safe(text):
    """Replace non-latin1 characters that Helvetica can't render."""
    if not isinstance(text, str):
        return str(text) if text is not None else '-'
    try:
        text.encode('latin-1')
        return text
    except UnicodeEncodeError:
        return text.encode('latin-1', errors='replace').decode('latin-1')


def write_pdf_report(filepath, old_label, new_label, cats, old_db, new_db):
    """Generate a clean PDF comparison report using fpdf2."""
    from fpdf import FPDF
    from datetime import datetime

    old_msg_count = len(old_db.messages)
    new_msg_count = len(new_db.messages)
    old_sig_count = sum(len(m.signals) for m in old_db.messages.values())
    new_sig_count = sum(len(m.signals) for m in new_db.messages.values())

    pdf = FPDF(orientation='L', unit='mm', format='A4')
    pdf.set_auto_page_break(auto=True, margin=15)

    # Colors
    hdr_r, hdr_g, hdr_b = 31, 78, 121  # #1F4E79
    alt_r, alt_g, alt_b = 242, 242, 242  # #F2F2F2
    diff_r, diff_g, diff_b = 255, 199, 206  # #FFC7CE

    # Margins: 10mm each side
    margin = 10
    page_height = 297  # A4 height used for all pages

    def add_title_page():
        pdf.add_page()
        pdf.set_font('Helvetica', 'B', 24)
        pdf.set_text_color(hdr_r, hdr_g, hdr_b)
        pdf.cell(0, 20, 'DBC Comparison Report', new_x="LMARGIN", new_y="NEXT")
        pdf.set_font('Helvetica', '', 12)
        pdf.set_text_color(120, 120, 120)
        pdf.cell(0, 8, f'{old_label}  vs  {new_label}', new_x="LMARGIN", new_y="NEXT")
        pdf.cell(0, 8, f'Generated: {datetime.now().strftime("%Y-%m-%d %H:%M")}', new_x="LMARGIN", new_y="NEXT")
        pdf.ln(10)

        # Summary cards
        pdf.set_font('Helvetica', 'B', 14)
        pdf.set_text_color(hdr_r, hdr_g, hdr_b)
        pdf.cell(0, 10, 'Summary', new_x="LMARGIN", new_y="NEXT")
        pdf.set_draw_color(214, 228, 240)
        pdf.line(margin, pdf.get_y(), 287, pdf.get_y())
        pdf.ln(4)

        metrics = [
            ('Messages (Base)', old_msg_count),
            ('Messages (New)', new_msg_count),
            ('Signals (Base)', old_sig_count),
            ('Signals (New)', new_sig_count),
            ('New Messages', len(cats['new_messages'])),
            ('Removed Messages', len(cats['removed_messages'])),
            ('Modified Messages', len(cats['modified_messages'])),
            ('New Signals', len(cats['new_signals'])),
            ('Removed Signals', len(cats['removed_signals'])),
            ('Modified Signals', len(cats['modified_signals'])),
        ]
        pdf.set_font('Helvetica', '', 11)
        for label, val in metrics:
            pdf.set_text_color(80, 80, 80)
            pdf.cell(80, 7, label)
            pdf.set_text_color(hdr_r, hdr_g, hdr_b)
            pdf.set_font('Helvetica', 'B', 11)
            pdf.cell(30, 7, str(val), new_x="LMARGIN", new_y="NEXT")
            pdf.set_font('Helvetica', '', 11)

    def add_table_section(title, headers, rows, col_widths, diff_col_sets=None):
        # Auto-expand page width to fit all columns
        table_width = sum(col_widths)
        needed_width = table_width + 2 * margin
        # Minimum landscape A4 width (297mm), expand if table needs more
        pw = max(297, needed_width)
        pdf.add_page(format=(pw, page_height))

        pdf.set_font('Helvetica', 'B', 14)
        pdf.set_text_color(hdr_r, hdr_g, hdr_b)
        pdf.cell(0, 10, title, new_x="LMARGIN", new_y="NEXT")
        pdf.set_draw_color(214, 228, 240)
        pdf.line(margin, pdf.get_y(), pw - margin, pdf.get_y())
        pdf.ln(4)

        if not rows:
            pdf.set_font('Helvetica', 'I', 10)
            pdf.set_text_color(150, 150, 150)
            pdf.cell(0, 8, f'No {title.lower()}', new_x="LMARGIN", new_y="NEXT")
            return

        # Scale font and row height based on available space
        font_sz = 9 if table_width < 500 else 8 if table_width < 800 else 7
        hdr_font_sz = font_sz + 1
        row_h = 7 if font_sz >= 9 else 6
        hdr_h = row_h + 1
        max_chars = 80 if table_width > 500 else 50

        # Table header
        pdf.set_font('Helvetica', 'B', hdr_font_sz)
        pdf.set_fill_color(hdr_r, hdr_g, hdr_b)
        pdf.set_text_color(255, 255, 255)
        for i, h in enumerate(headers):
            pdf.cell(col_widths[i], hdr_h, _pdf_safe(h)[:30], border=1, fill=True, align='C')
        pdf.ln()

        # Table rows
        pdf.set_font('Helvetica', '', font_sz)
        page_break_y = page_height - 20
        for r_idx, row in enumerate(rows):
            diffs = diff_col_sets[r_idx] if diff_col_sets else set()
            for c_idx, val in enumerate(row):
                if c_idx in diffs:
                    pdf.set_fill_color(diff_r, diff_g, diff_b)
                    fill = True
                elif r_idx % 2 == 1:
                    pdf.set_fill_color(alt_r, alt_g, alt_b)
                    fill = True
                else:
                    fill = False
                pdf.set_text_color(50, 50, 50)
                txt = _pdf_safe(val) if val is not None and val != '' else '-'
                txt = txt.replace('\n', ', ')[:max_chars]
                pdf.cell(col_widths[c_idx], row_h, txt, border=1, fill=fill)
            pdf.ln()
            if pdf.get_y() > page_break_y:
                pdf.add_page(format=(pw, page_height))
                # Re-draw header
                pdf.set_font('Helvetica', 'B', hdr_font_sz)
                pdf.set_fill_color(hdr_r, hdr_g, hdr_b)
                pdf.set_text_color(255, 255, 255)
                for i, h in enumerate(headers):
                    pdf.cell(col_widths[i], hdr_h, _pdf_safe(h)[:30], border=1, fill=True, align='C')
                pdf.ln()
                pdf.set_font('Helvetica', '', font_sz)

    add_title_page()

    # Message col widths (8 cols, total ~1169mm for landscape A0)
    msg_w = [50, 160, 50, 200, 60, 50, 100, 200]
    sig_w = [50, 130, 130, 45, 40, 45, 50, 50, 45, 45, 170, 90]

    # New Messages
    new_msg_rows = [build_msg_summary_row(m) for m in cats['new_messages']]
    add_table_section("New Messages", CAT_MSG_HEADERS, new_msg_rows, msg_w)

    # Removed Messages
    rem_msg_rows = [build_msg_summary_row(m) for m in cats['removed_messages']]
    add_table_section("Removed Messages", CAT_MSG_HEADERS, rem_msg_rows, msg_w)

    # Modified Messages
    if cats['modified_messages']:
        mod_hdrs = []
        for h in CAT_MSG_HEADERS:
            mod_hdrs.extend([f"{h} (Old)", f"{h} (New)"])
        mod_rows = []
        mod_diffs = []
        mod_w = [w2 for w in msg_w for w2 in [w // 2, w // 2]]
        for old_msg, new_msg, diff_indices in cats['modified_messages']:
            old_r = build_msg_summary_row(old_msg)
            new_r = build_msg_summary_row(new_msg)
            row = []
            ds = set()
            for i in range(len(old_r)):
                row.extend([old_r[i], new_r[i]])
                if i in diff_indices:
                    ds.add(i * 2)
                    ds.add(i * 2 + 1)
            mod_rows.append(row)
            mod_diffs.append(ds)
        add_table_section("Modified Messages", mod_hdrs, mod_rows, mod_w, mod_diffs)
    else:
        add_table_section("Modified Messages", CAT_MSG_HEADERS, [], msg_w)

    # New Signals
    new_sig_rows = [build_sig_summary_row(m, s) for m, s in cats['new_signals']]
    add_table_section("New Signals", CAT_SIG_HEADERS, new_sig_rows, sig_w)

    # Removed Signals
    rem_sig_rows = [build_sig_summary_row(m, s) for m, s in cats['removed_signals']]
    add_table_section("Removed Signals", CAT_SIG_HEADERS, rem_sig_rows, sig_w)

    # Modified Signals
    if cats['modified_signals']:
        fixed_hdrs = ["Msg ID", "Msg Name", "Signal"]
        paired_hdrs = list(fixed_hdrs)
        for h in MOD_SIG_PAIRED_HEADERS:
            paired_hdrs.extend([f"{h} (Old)", f"{h} (New)"])
        ms_w = [50, 100, 100] + [40, 40] * len(MOD_SIG_PAIRED_HEADERS)
        ms_rows = []
        ms_diffs = []
        for old_msg, new_msg, old_sig, new_sig, diff_indices in cats['modified_signals']:
            old_sr = build_sig_summary_row(old_msg, old_sig)
            new_sr = build_sig_summary_row(new_msg, new_sig)
            row = [old_sr[0], old_sr[1], old_sr[2]]
            ds = set()
            for i in range(len(MOD_SIG_PAIRED_HEADERS)):
                row.extend([old_sr[3 + i], new_sr[3 + i]])
                if i in diff_indices:
                    ds.add(3 + i * 2)
                    ds.add(3 + i * 2 + 1)
            ms_rows.append(row)
            ms_diffs.append(ds)
        add_table_section("Modified Signals", paired_hdrs, ms_rows, ms_w, ms_diffs)
    else:
        add_table_section("Modified Signals", CAT_SIG_HEADERS[:3], [], [18, 40, 40])

    pdf.output(filepath)
    print(f"  Saved PDF: {filepath}")


# ==============================================================================
# Bus Matching
# ==============================================================================

def extract_bus_prefix(filename):
    """Extract bus prefix like '01_BusName' from a filename."""
    m = re.match(r'(\d+_[\w]+?)_', filename, re.IGNORECASE)
    if m:
        return m.group(1)
    return None


def extract_bus_name(bus_prefix):
    """Extract bus name from prefix, e.g. 'BusName' from '01_BusName'."""
    parts = bus_prefix.split('_', 1)
    return parts[1] if len(parts) == 2 else bus_prefix


def find_dbc_files(folder):
    """Find all .dbc files in a folder, indexed by bus prefix."""
    dbc_files = {}
    for fname in os.listdir(folder):
        if fname.lower().endswith('.dbc') and not fname.startswith('~$'):
            prefix = extract_bus_prefix(fname)
            if prefix:
                dbc_files[prefix] = os.path.join(folder, fname)
    return dbc_files


# ==============================================================================
# Main
# ==============================================================================

def main():
    parser = argparse.ArgumentParser(description='Compare DBC files between two EP versions')
    parser.add_argument('old_folder', help='Path to old version folder')
    parser.add_argument('new_folder', help='Path to new version folder')
    parser.add_argument('--output', '-o', help='Output folder name (auto-generated if not specified)')
    args = parser.parse_args()

    old_folder = args.old_folder.rstrip('/')
    new_folder = args.new_folder.rstrip('/')
    old_name = os.path.basename(old_folder)
    new_name = os.path.basename(new_folder)

    if args.output:
        output_dir = args.output
    else:
        output_dir = os.path.join(
            os.path.dirname(old_folder),
            f"DBC_Compare_{old_name}_vs_{new_name}"
        )

    os.makedirs(output_dir, exist_ok=True)
    print(f"Output directory: {output_dir}")
    print(f"Columns per side: {NUM_COLS_PER_SIDE} ({NUM_MSG_COLS} msg + {NUM_SIG_COLS} sig)")

    old_dbc_files = find_dbc_files(old_folder)
    new_dbc_files = find_dbc_files(new_folder)

    print(f"\nOld folder ({old_name}): {len(old_dbc_files)} DBC files")
    for prefix, path in sorted(old_dbc_files.items()):
        print(f"  {prefix}: {os.path.basename(path)}")

    print(f"\nNew folder ({new_name}): {len(new_dbc_files)} DBC files")
    for prefix, path in sorted(new_dbc_files.items()):
        print(f"  {prefix}: {os.path.basename(path)}")

    all_buses = sorted(set(list(old_dbc_files.keys()) + list(new_dbc_files.keys())))

    print(f"\n{'='*60}")
    print(f"Comparing {old_name} vs {new_name}")
    print(f"{'='*60}")

    summary = []

    for bus_prefix in all_buses:
        bus_name = extract_bus_name(bus_prefix)
        old_path = old_dbc_files.get(bus_prefix)
        new_path = new_dbc_files.get(bus_prefix)

        if not old_path and not new_path:
            continue
        if not old_path:
            print(f"\n[{bus_name}] Only in new version: {os.path.basename(new_path)}")
            continue
        if not new_path:
            print(f"\n[{bus_name}] Only in old version: {os.path.basename(old_path)}")
            continue

        print(f"\n[{bus_name}] Comparing...")
        print(f"  Old: {os.path.basename(old_path)}")
        print(f"  New: {os.path.basename(new_path)}")

        old_db = parse_dbc(old_path)
        new_db = parse_dbc(new_path)

        comparison_rows = compare_dbc_files(old_db, new_db)

        diff_count = sum(1 for _, _, has_diff, _ in comparison_rows if has_diff)
        total_rows = len(comparison_rows)
        print(f"  Total signal rows: {total_rows}")
        print(f"  Rows with differences: {diff_count}")

        old_rel = f"{old_name}\\{os.path.basename(old_path)}"
        new_rel = f"{new_name}\\{os.path.basename(new_path)}"

        base_name = f"DBC_Compare_{old_name}_vs_{new_name}_{bus_name}"
        output_path = os.path.join(output_dir, f"{base_name}.xlsx")
        write_comparison_xlsx(output_path, old_rel, new_rel, comparison_rows,
                              old_db=old_db, new_db=new_db,
                              old_label=old_name, new_label=new_name)

        # Generate HTML and PDF reports
        cats = categorize_changes(old_db, new_db)
        write_html_report(
            os.path.join(output_dir, f"{base_name}.html"),
            old_name, new_name, cats, old_db, new_db)
        write_pdf_report(
            os.path.join(output_dir, f"{base_name}.pdf"),
            old_name, new_name, cats, old_db, new_db)

        summary.append((bus_name, total_rows, diff_count))

    # Print summary
    print(f"\n{'='*60}")
    print("SUMMARY")
    print(f"{'='*60}")
    print(f"{'Bus':<15} {'Total Rows':>12} {'Differences':>13}")
    print(f"{'-'*15} {'-'*12} {'-'*13}")
    total_all = 0
    diff_all = 0
    for bus_name, total, diffs in summary:
        print(f"{bus_name:<15} {total:>12} {diffs:>13}")
        total_all += total
        diff_all += diffs
    print(f"{'-'*15} {'-'*12} {'-'*13}")
    print(f"{'TOTAL':<15} {total_all:>12} {diff_all:>13}")

    print(f"\nAll columns per side:")
    print(f"  Message: {', '.join(MSG_HEADERS)}")
    print(f"  Signal:  {', '.join(SIG_HEADERS)}")
    print(f"\nOutput saved to: {output_dir}")


if __name__ == "__main__":
    main()
